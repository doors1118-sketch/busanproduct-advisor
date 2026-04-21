"""국가·정부공공기관 지역업체 계약률 순위 엑셀 생성"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

wb = openpyxl.Workbook()

# ── 스타일 ──
hdr_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
hdr_fill_top = PatternFill("solid", fgColor="2E75B6")   # 상위 파랑
hdr_fill_btm = PatternFill("solid", fgColor="C0504D")   # 하위 빨강
body_font = Font(name="맑은 고딕", size=10)
title_font = Font(name="맑은 고딕", size=14, bold=True)
rank_font = Font(name="맑은 고딕", size=10, bold=True)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
right_align = Alignment(horizontal="right", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
pct_fmt = '0.0%'
num_fmt = '#,##0'

# ── 데이터 ──
top15 = [
    (1, "한국토지주택공사", 2900000000, 2900000000, 0.991),
    (2, "한국전력공사", 1100000000, 1000000000, 0.941),
    (3, "해양경찰청", 13000000000, 11200000000, 0.864),
    (4, "한국수산자원공단", 7000000000, 5800000000, 0.827),
    (5, "관세청", 1100000000, 900000000, 0.800),
    (6, "법무부", 1900000000, 1500000000, 0.792),
    (7, "한국해양수산연수원", 6800000000, 5400000000, 0.786),
    (8, "한국공항공사", 2400000000, 1800000000, 0.775),
    (9, "국립부산과학관", 2100000000, 1500000000, 0.713),
    (10, "동의대학교", 3200000000, 2200000000, 0.673),
    (11, "부산대학교", 64500000000, 41900000000, 0.651),
    (12, "식품의약품안전처", 1600000000, 1000000000, 0.631),
    (13, "경찰청", 1000000000, 600000000, 0.621),
    (14, "부경대학교", 21300000000, 12700000000, 0.596),
    (15, "한국해양대학교", 9800000000, 5600000000, 0.576),
]

btm15 = [
    (1, "주택도시보증공사", 13400000000, 42000000, 0.003),
    (2, "한국주택금융공사", 19500000000, 99400000, 0.005),
    (3, "한국수력원자력(주)고리원자력본부", 2800000000, 90830000, 0.032),
    (4, "한국원자력환경복원연구원", 17900000000, 1100000000, 0.061),
    (5, "과학기술정보통신부", 3600000000, 300000000, 0.072),
    (6, "국토교통부", 16900000000, 1300000000, 0.075),
    (7, "게임물관리위원회", 1700000000, 100000000, 0.082),
    (8, "영화진흥위원회", 9400000000, 1100000000, 0.117),
    (9, "해양수산부", 438300000000, 83900000000, 0.191),
    (10, "기술보증기금", 1500000000, 300000000, 0.209),
    (11, "국세청", 11300000000, 2500000000, 0.224),
    (12, "해양환경공단", 2200000000, 500000000, 0.235),
    (13, "신라대학교", 1100000000, 300000000, 0.265),
    (14, "한국해양진흥공사", 3000000000, 800000000, 0.274),
    (15, "한국해양과학기술원", 7800000000, 2600000000, 0.340),
]

headers = ["순위", "수요기관명", "총 발주액", "지역업체 수주액", "수주율"]
col_widths = [8, 35, 18, 18, 12]

def write_section(ws, start_row, title, data, hdr_fill):
    # 제목 행
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = title_font
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # 헤더
    hr = start_row + 1
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = border

    # 데이터
    for ri, (rank, name, total, local, rate) in enumerate(data, hr + 1):
        ws.cell(row=ri, column=1, value=rank).font = rank_font
        ws.cell(row=ri, column=1).alignment = center
        ws.cell(row=ri, column=2, value=name).font = body_font
        ws.cell(row=ri, column=2).alignment = left_align

        tc = ws.cell(row=ri, column=3, value=total)
        tc.font = body_font
        tc.number_format = '#,##0'
        tc.alignment = right_align

        lc = ws.cell(row=ri, column=4, value=local)
        lc.font = body_font
        lc.number_format = '#,##0'
        lc.alignment = right_align

        rc = ws.cell(row=ri, column=5, value=rate)
        rc.font = body_font
        rc.number_format = pct_fmt
        rc.alignment = center

        for col in range(1, 6):
            ws.cell(row=ri, column=col).border = border

    return hr + len(data) + 1  # next available row


ws = wb.active
ws.title = "지역업체 계약률"

# 열 너비
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# 상위 15
next_row = write_section(ws, 1, "🔝 상위 15개 기관 (수주율 높은 순)", top15, hdr_fill_top)

# 빈 행
next_row += 1

# 하위 15
write_section(ws, next_row, "🔻 하위 15개 기관 (수주율 낮은 순)", btm15, hdr_fill_btm)

out = r"C:\Users\COMTREE\Desktop\메뉴얼 제작\국가공공기관_지역업체_계약률_순위.xlsx"
wb.save(out)
print(f"저장 완료: {out}")
